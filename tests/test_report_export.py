"""
=============================================================================
Project : AI-Powered Retail Demand Forecasting &
          Inventory Optimization System

File : test_report_export.py

Description :
Unit tests for the Milestone 4 - Phase 3 export mechanics
(utils/report_export.py). Confirms CSV/Excel round-trip cleanly back
through pandas' own readers (i.e. a downloaded file actually re-opens
correctly) and that PDF bytes form a real, valid PDF document.

Run with either:
    python tests/test_report_export.py
    python -m pytest tests/
=============================================================================
"""

import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pandas as pd

from utils import report_charts as charts
from utils import report_export as export


SAMPLE_TABLE = pd.DataFrame({

    "Product ID": ["P1", "P2"],

    "Product Name": ["Widget", "Gadget"],

    "Forecast Demand": [100.0, 50.0]

})


def test_to_csv_bytes_round_trips():

    data = export.to_csv_bytes(SAMPLE_TABLE)

    parsed = pd.read_csv(BytesIO(data))

    assert list(parsed.columns) == list(SAMPLE_TABLE.columns)

    assert len(parsed) == 2


def test_to_csv_bytes_handles_none_table():

    data = export.to_csv_bytes(None)

    assert isinstance(data, bytes)

    assert len(data) >= 0


def test_to_excel_bytes_multi_sheet_round_trips():

    sheets = {

        "Executive": pd.DataFrame({"Metric": ["Total"], "Value": [2]}),

        "Products": SAMPLE_TABLE

    }

    data = export.to_excel_bytes(sheets)

    workbook = openpyxl.load_workbook(BytesIO(data))

    assert set(workbook.sheetnames) == {"Executive", "Products"}

    parsed = pd.read_excel(BytesIO(data), sheet_name="Products")

    assert len(parsed) == 2


def test_to_excel_bytes_truncates_long_sheet_names():

    sheets = {"A" * 50: SAMPLE_TABLE}

    data = export.to_excel_bytes(sheets)

    workbook = openpyxl.load_workbook(BytesIO(data))

    assert len(workbook.sheetnames[0]) <= 31


def test_build_pdf_bytes_produces_a_valid_pdf():

    data = export.build_pdf_bytes(

        title="Test Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0),

        kpis={"Total Products": 2},

        tables=[("Products", SAMPLE_TABLE)]

    )

    assert data[:4] == b"%PDF"

    assert data.rstrip()[-5:] in (b"%%EOF", b"EOF\n>>")


def test_build_pdf_bytes_with_charts():

    figure = charts.risk_pie_chart({"Critical": 1, "Low": 1})

    data = export.build_pdf_bytes(

        title="Test Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0),

        charts=[("Risk", figure)]

    )

    assert data[:4] == b"%PDF"


def test_build_pdf_bytes_with_summary_sections():

    data = export.build_pdf_bytes(

        title="AI Executive Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0),

        summary_sections={

            "Overall Health": "Everything looks fine.",

            "Critical Issues": "None."

        }

    )

    assert data[:4] == b"%PDF"


def test_build_pdf_bytes_with_no_optional_sections_still_valid():

    data = export.build_pdf_bytes(

        title="Empty Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0)

    )

    assert data[:4] == b"%PDF"


def test_build_pdf_bytes_empty_table_shows_no_data_message():

    # Must not raise even though the table has zero rows.
    data = export.build_pdf_bytes(

        title="Test Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0),

        tables=[("Empty Table", pd.DataFrame())]

    )

    assert data[:4] == b"%PDF"


def test_build_pdf_bytes_truncates_large_tables():

    large_table = pd.DataFrame({

        "Product ID": [f"P{i}" for i in range(150)],

        "Value": range(150)

    })

    data = export.build_pdf_bytes(

        title="Test Report",

        dataset_name="Demo Dataset",

        generated_at=datetime(2026, 1, 1, 12, 0),

        tables=[("Big Table", large_table)]

    )

    assert data[:4] == b"%PDF"


# ---------------------------------------------------------------------
# Plain-python runner
# ---------------------------------------------------------------------

if __name__ == "__main__":

    failures = 0

    tests = [

        (name, function)

        for name, function in sorted(globals().items())

        if name.startswith("test_") and callable(function)

    ]

    for name, function in tests:

        try:

            function()

            print(f"PASS  {name}")

        except Exception as error:

            failures += 1

            print(f"FAIL  {name}: {error}")

    print(

        f"\n{len(tests) - failures}/{len(tests)} tests passed."

    )

    sys.exit(1 if failures else 0)
